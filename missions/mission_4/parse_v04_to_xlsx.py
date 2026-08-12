"""
v0.4 → v0.5 转换器
解析 v0.4 phase 文件, 抽取 18 字段, 写入 mission_spec.xlsx
"""
import re
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ========== 配置 ==========
V04_DIR = Path('/workspace/missions/mission_4')
PHASE_FILES = [
    ('P1', V04_DIR / 'v0.4_component_content_p1.md'),
    ('P2', V04_DIR / 'v0.4_component_content_p2.md'),
    ('P3', V04_DIR / 'v0.4_component_content_p3.md'),
    ('P4', V04_DIR / 'v0.4_component_content_p4.md'),
]
OUTPUT_XLSX = V04_DIR / 'v0.5_mission_spec.xlsx'

# 18 字段列定义
COLUMNS = [
    'Phase',                              # 1
    'script_step',                        # 2
    'Screen Display Category',            # 3
    'Display Format',                     # 4
    'Component 序号',                     # 5
    'Display Text',                       # 6
    'Display Image',                      # 7
    'Video Play',                         # 8
    'Audio - 字幕',                       # 9
    'Audio - 音色',                       # 10
    'Interactive Flow',                   # 11
    'Kai Script 1',                       # 12 (trigger + script)
    'Kai Script 2',                       # 13 (trigger + script)
    'Possible Student Response',          # 14 (trigger + feedback)
    'Kai Feedback (Correct)',             # 15
    'Kai Feedback (Wrong)',               # 16
    'Transition Script',                  # 17 (trigger + script)
    'Knowledge Point',                    # 18
]

# Knowledge Point 子字段顺序
KP_FIELDS = ['word', 'grammar', 'phrase', 'pattern', 'socialExpression', 'pinyin']


# ========== 解析函数 ==========
def parse_lesson_step(section_text):
    """解析一个 lesson_step section, 返回 18 字段字典"""
    data = {}

    # 1. lesson_step 编号 (e.g. "1.1") + Component 序号
    m = re.search(r'## lesson_step (\d+\.\d+): (CMP-\d+)\s+(.+)', section_text)
    if m:
        ls_id, comp_id, comp_name = m.group(1), m.group(2), m.group(3)
        data['_ls_id'] = ls_id
        data['Component 序号'] = comp_id

    # 2. Display Text (代码块)
    m = re.search(r'### Display Text\s*\n\n```\n(.*?)\n```', section_text, re.DOTALL)
    if m:
        data['Display Text'] = m.group(1).strip()

    # 3. 其他 17 字段 (markdown 表格)
    # 找 "### 其他 17 字段" 之后的表格
    table_match = re.search(
        r'### 其他 17 字段\s*\n\n\| 字段 \| 值 \|\s*\n\|---+\|---\|\s*\n(.*?)(?=\n\n---|\n\n##|\Z)',
        section_text, re.DOTALL
    )
    if table_match:
        table_rows = table_match.group(1).strip().split('\n')
        for row in table_rows:
            row_m = re.match(r'\| \*\*(.+?)\*\* \| (.+?) \|$', row)
            if not row_m:
                continue
            field = row_m.group(1).strip()
            value = row_m.group(2).strip()

            # Phase (从 lesson_step 标题的 Phase 标注或上一个 ## 标题取)
            # 这里 Phase 由 phase 文件决定 (P1/P2/P3/P4)

            # script_step
            if field == 'Phase':
                data['Phase'] = value
            elif field == 'script_step':
                data['script_step'] = int(value) if value.isdigit() else value
            elif field == 'Screen Display Category':
                data['Screen Display Category'] = value
            elif field == 'Display Format':
                data['Display Format'] = value
            elif field == 'Component 序号':
                pass  # 已从标题取
            elif field == 'Display Text':
                pass  # 已从代码块取
            elif field == 'Display Image':
                data['Display Image'] = value
            elif field == 'Video Play':
                data['Video Play'] = value
            elif field == 'Audio - 字幕':
                data['Audio - 字幕'] = value
            elif field == 'Audio - 音色':
                data['Audio - 音色'] = value
            elif field == 'Interactive Flow':
                data['Interactive Flow'] = value
            elif field.startswith('Kai Script 1'):
                # Kai Script 1.trigger / Kai Script 1.script (合并)
                if 'trigger' in field:
                    data['_kai1_trigger'] = value
                else:
                    data['_kai1_script'] = value
            elif field.startswith('Kai Script 2'):
                if 'trigger' in field:
                    data['_kai2_trigger'] = value
                else:
                    data['_kai2_script'] = value
            elif field.startswith('Possible Student Response'):
                if 'trigger' in field:
                    data['_resp_trigger'] = value
                else:
                    data['_resp_feedback'] = value
            elif field == 'Kai Feedback (Correct)':
                data['Kai Feedback (Correct)'] = value
            elif field == 'Kai Feedback (Wrong)':
                data['Kai Feedback (Wrong)'] = value
            elif field.startswith('Transition Script'):
                if 'trigger' in field:
                    data['_trans_trigger'] = value
                else:
                    data['_trans_script'] = value
            elif field == 'Knowledge Point':
                data['Knowledge Point'] = value

    return data


def merge_complex_fields(data):
    """合并 trigger + script 字段为单列"""
    # Kai Script 1
    if '_kai1_trigger' in data or '_kai1_script' in data:
        t = data.get('_kai1_trigger', '—')
        s = data.get('_kai1_script', '—')
        if t == '—' and s == '—':
            data['Kai Script 1'] = '—'
        else:
            data['Kai Script 1'] = f"trigger: {t}\nscript: {s}"

    # Kai Script 2
    if '_kai2_trigger' in data or '_kai2_script' in data:
        t = data.get('_kai2_trigger', '—')
        s = data.get('_kai2_script', '—')
        if t == '—' and s == '—':
            data['Kai Script 2'] = '—'
        else:
            data['Kai Script 2'] = f"trigger: {t}\nscript: {s}"

    # Possible Student Response
    if '_resp_trigger' in data or '_resp_feedback' in data:
        t = data.get('_resp_trigger', '—')
        f = data.get('_resp_feedback', '—')
        if t == '—' and f == '—':
            data['Possible Student Response'] = '—'
        else:
            data['Possible Student Response'] = f"trigger: {t}\nfeedback: {f}"

    # Transition Script
    if '_trans_trigger' in data or '_trans_script' in data:
        t = data.get('_trans_trigger', '—')
        s = data.get('_trans_script', '—')
        if t == '—' and s == '—':
            data['Transition Script'] = '—'
        else:
            data['Transition Script'] = f"trigger: {t}\nscript: {s}"

    # 清理临时字段
    for k in ['_kai1_trigger', '_kai1_script', '_kai2_trigger', '_kai2_script',
              '_resp_trigger', '_resp_feedback', '_trans_trigger', '_trans_script', '_ls_id']:
        data.pop(k, None)

    return data


def parse_knowledge_point(value):
    """将 Knowledge Point 字符串格式化为易读形式"""
    if not value or value == '—':
        return '—'
    # value 格式: { word: "...", pattern: "..." } 等
    # 提取每对 key: value
    parts = re.findall(r'(\w+):\s*([^,}]+?)(?=[,}]|$)', value)
    if not parts:
        return value
    lines = []
    for k, v in parts:
        v = v.strip().strip('"').strip("'")
        if v and v != '—':
            lines.append(f"{k}: {v}")
    return '\n'.join(lines) if lines else '—'


def fill_defaults(data):
    """为缺失字段填默认值"""
    for col in COLUMNS:
        if col not in data:
            data[col] = '—'
        if data[col] is None:
            data[col] = '—'
    return data


# ========== 主流程 ==========
def main():
    all_rows = []

    for phase, file_path in PHASE_FILES:
        if not file_path.exists():
            print(f"  ⚠️ 文件不存在: {file_path}")
            continue
        with open(file_path) as f:
            content = f.read()

        # 按 ## lesson_step 分割
        sections = re.split(r'\n## lesson_step ', content)
        for sec in sections[1:]:  # 跳过第一个 (无 lesson_step)
            # 加上 ## 前缀以便解析
            data = parse_lesson_step('## lesson_step ' + sec)
            if not data.get('Component 序号'):
                continue

            # 强制 Phase 从文件名取 (避免源文档里可能缺失)
            if 'Phase' not in data or data['Phase'] == '—':
                data['Phase'] = phase

            data = merge_complex_fields(data)
            data['Knowledge Point'] = parse_knowledge_point(data.get('Knowledge Point', '—'))
            data = fill_defaults(data)
            all_rows.append(data)

    print(f"  ✅ 共解析 {len(all_rows)} 个 mission step")

    # 写入 xlsx
    wb = Workbook()
    ws = wb.active
    ws.title = 'mission_spec'

    # 表头
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 数据行
    for row_idx, row_data in enumerate(all_rows, 2):
        for col_idx, col_name in enumerate(COLUMNS, 1):
            value = row_data.get(col_name, '—')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    # 列宽
    col_widths = {
        'Phase': 8,
        'script_step': 10,
        'Screen Display Category': 22,
        'Display Format': 18,
        'Component 序号': 12,
        'Display Text': 50,
        'Display Image': 35,
        'Video Play': 35,
        'Audio - 字幕': 25,
        'Audio - 音色': 25,
        'Interactive Flow': 50,
        'Kai Script 1': 45,
        'Kai Script 2': 40,
        'Possible Student Response': 40,
        'Kai Feedback (Correct)': 35,
        'Kai Feedback (Wrong)': 30,
        'Transition Script': 35,
        'Knowledge Point': 35,
    }
    for col_idx, col_name in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 20)

    # 冻结表头
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT_XLSX)
    print(f"  ✅ 写入 {OUTPUT_XLSX}")
    print(f"  文件大小: {OUTPUT_XLSX.stat().st_size:,} bytes")


if __name__ == '__main__':
    main()
